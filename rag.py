"""RAG — index and search local files via Qdrant (hybrid dense + sparse)."""

import os, time, uuid, threading
from pathlib import Path
import config, db, logger

_log = logger.get("rag")

# Separate Qdrant collection for RAG (not mixed with agent memory)
RAG_COLLECTION = "qwe_rag"
CHUNK_SIZE = 800        # chars (~200 tokens), configurable via settings
CHUNK_OVERLAP = 100     # chars (~25 tokens), configurable via settings
SUPPORTED_EXTENSIONS = {
    # Plain text & markdown
    ".txt", ".md", ".markdown", ".rst", ".asciidoc", ".adoc", ".tex", ".log",
    # Data / config
    ".json", ".jsonc", ".ndjson", ".csv", ".tsv",
    ".yaml", ".yml", ".toml", ".ini", ".cfg", ".conf", ".env", ".properties",
    ".xml", ".xsd", ".svg",
    # Web
    ".html", ".htm", ".xhtml", ".css", ".scss", ".sass", ".less",
    # Code
    ".py", ".pyi", ".ipynb",
    ".js", ".jsx", ".mjs", ".cjs", ".ts", ".tsx",
    ".go", ".rs", ".java", ".kt", ".kts", ".scala", ".groovy",
    ".c", ".cpp", ".cc", ".cxx", ".h", ".hpp", ".hxx", ".m", ".mm",
    ".rb", ".php", ".pl", ".pm", ".lua", ".r", ".jl", ".dart",
    ".swift", ".sql", ".graphql", ".gql", ".proto", ".thrift",
    ".sh", ".bash", ".zsh", ".fish", ".ps1", ".bat", ".cmd",
    ".vim", ".nix", ".tf", ".tfvars", ".hcl",
    # Docker / CI
    ".dockerfile", ".containerfile",
}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}
PDF_EXTENSION = ".pdf"
# Office & book formats — parsed by specialist readers in _read_file
OFFICE_EXTENSIONS = {".docx", ".pptx", ".xlsx", ".odt", ".epub", ".rtf"}
URL_SCHEMES = ("http://", "https://")
ALL_INDEXABLE = SUPPORTED_EXTENSIONS | IMAGE_EXTENSIONS | {PDF_EXTENSION} | OFFICE_EXTENSIONS
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
VISION_RATE_LIMIT = 1.0  # seconds between vision API calls
SCANNED_PAGE_THRESHOLD = 50  # chars — below this, page is likely scanned
MAX_SCAN_FILES = 1000

# Lazy imports
_qclient = None


def _get_qdrant():
    """Get shared Qdrant client from memory module (avoids duplicate connections)."""
    global _qclient
    if _qclient is None:
        import memory
        _qclient = memory._get_qdrant()
        # Ensure RAG collection exists with v2 schema (named dense + sparse)
        from qdrant_client.models import (
            VectorParams, Distance, PayloadSchemaType, Datatype,
            SparseVectorParams,
        )
        cols = [c.name for c in _qclient.get_collections().collections]
        if RAG_COLLECTION not in cols:
            _qclient.create_collection(
                RAG_COLLECTION,
                vectors_config={
                    "dense": VectorParams(
                        size=memory.EMBED_DIM,
                        distance=Distance.COSINE,
                        datatype=Datatype.FLOAT16,
                    ),
                },
                sparse_vectors_config={
                    "sparse": SparseVectorParams(),
                },
            )
        # Ensure payload indexes
        try:
            info = _qclient.get_collection(RAG_COLLECTION)
            existing = set(info.payload_schema.keys()) if info.payload_schema else set()
            if "file_path" not in existing:
                _qclient.create_payload_index(RAG_COLLECTION, "file_path", PayloadSchemaType.KEYWORD)
                _log.info("created payload index: file_path")
            if "tags" not in existing:
                _qclient.create_payload_index(RAG_COLLECTION, "tags", PayloadSchemaType.KEYWORD)
                _log.info("created payload index: tags")
        except Exception as e:
            _log.debug(f"RAG payload index creation skipped: {e}")
    return _qclient


def _embed(text: str) -> list[float]:
    """Generate dense embedding via memory module's FastEmbed model."""
    import memory
    return memory._embed(text)


def _chunk_text(text: str) -> list[str]:
    """Split text into overlapping chunks. Uses configurable sizes from settings."""
    chunk_size = config.get("rag_chunk_size") if "rag_chunk_size" in config.EDITABLE_SETTINGS else CHUNK_SIZE
    chunk_overlap = config.get("rag_chunk_overlap") if "rag_chunk_overlap" in config.EDITABLE_SETTINGS else CHUNK_OVERLAP
    if len(text) <= chunk_size:
        return [text]
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        if chunk.strip():
            chunks.append(chunk)
        start += chunk_size - chunk_overlap
    return chunks


def _get_markitdown():
    """Lazy-load a MarkItDown instance. Cached on first call.

    Returns None if the package is missing (graceful fallback to stdlib readers).
    """
    global _markitdown_cache
    try:
        return _markitdown_cache
    except NameError:
        pass
    try:
        from markitdown import MarkItDown
        _markitdown_cache = MarkItDown(enable_plugins=False)
        _log.info("markitdown loaded — advanced document conversion enabled")
    except ImportError:
        _log.warning("markitdown not installed — falling back to stdlib readers")
        _markitdown_cache = None
    except Exception as e:
        _log.warning(f"markitdown init failed: {e} — using stdlib fallbacks")
        _markitdown_cache = None
    return _markitdown_cache


def _markitdown_convert(source) -> str | None:
    """Try to extract markdown via MarkItDown. Returns None if unavailable or failed.

    ``source`` may be a filesystem path (str / Path) or an http(s) URL.
    """
    md = _get_markitdown()
    if md is None:
        return None
    try:
        src = str(source)
        result = md.convert(src)
        text = getattr(result, "text_content", None) or getattr(result, "markdown", None)
        if text and text.strip():
            return text.strip()
    except Exception as e:
        _log.debug(f"markitdown failed on {source}: {e}")
    return None


def _strip_html(html: str) -> str:
    """Quick stdlib HTML → text. Drops script/style, collapses whitespace."""
    import re as _re
    # Remove script/style blocks
    html = _re.sub(r"<(script|style)[^>]*>.*?</\1>", "", html, flags=_re.S | _re.I)
    # Replace common block-level tags with newlines
    html = _re.sub(r"</(p|div|li|tr|h[1-6]|br|section|article|header|footer|nav|aside)[^>]*>", "\n", html, flags=_re.I)
    html = _re.sub(r"<br\s*/?>", "\n", html, flags=_re.I)
    # Strip all remaining tags
    html = _re.sub(r"<[^>]+>", "", html)
    # Decode common entities (minimal)
    from html import unescape
    html = unescape(html)
    # Collapse whitespace
    html = _re.sub(r"\n{3,}", "\n\n", html)
    html = _re.sub(r"[ \t]+", " ", html)
    return html.strip()


def _read_docx(path: Path) -> str | None:
    """Extract text from .docx (Word) via stdlib zipfile — no deps needed."""
    import zipfile
    try:
        with zipfile.ZipFile(str(path)) as z:
            xml = z.read("word/document.xml").decode("utf-8", errors="replace")
    except Exception as e:
        _log.error(f"docx read failed: {path}: {e}")
        return None
    # Join <w:t> runs with spaces; paragraphs (<w:p>) already separated
    import re as _re
    # Insert paragraph breaks
    xml = _re.sub(r"</w:p>", "\n", xml)
    # Tab stops
    xml = _re.sub(r"<w:tab[^/>]*/?>", "\t", xml)
    # Extract text runs
    parts = _re.findall(r"<w:t[^>]*>([^<]*)</w:t>", xml)
    return "".join(parts).strip() or None


def _read_pptx(path: Path) -> str | None:
    """Extract text from .pptx (PowerPoint) — iterate slide XMLs."""
    import zipfile, re as _re
    try:
        with zipfile.ZipFile(str(path)) as z:
            slide_names = sorted(n for n in z.namelist() if n.startswith("ppt/slides/slide") and n.endswith(".xml"))
            chunks = []
            for name in slide_names:
                xml = z.read(name).decode("utf-8", errors="replace")
                # PowerPoint uses <a:t> for text runs
                parts = _re.findall(r"<a:t[^>]*>([^<]*)</a:t>", xml)
                if parts:
                    chunks.append("\n".join(parts))
            return "\n\n---\n\n".join(chunks).strip() or None
    except Exception as e:
        _log.error(f"pptx read failed: {path}: {e}")
        return None


def _read_xlsx(path: Path) -> str | None:
    """Extract text from .xlsx — prefer openpyxl, fall back to stdlib."""
    # Try openpyxl first (richer output, handles formulas)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), data_only=True, read_only=True)
        out = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            out.append(f"# Sheet: {sheet}")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    out.append("\t".join(cells))
        return "\n".join(out).strip() or None
    except ImportError:
        pass
    except Exception as e:
        _log.error(f"xlsx read failed (openpyxl): {path}: {e}")
        # Fall through to stdlib
    # Stdlib fallback: unpack sharedStrings.xml + sheet cells
    import zipfile, re as _re
    try:
        with zipfile.ZipFile(str(path)) as z:
            shared = []
            if "xl/sharedStrings.xml" in z.namelist():
                sx = z.read("xl/sharedStrings.xml").decode("utf-8", errors="replace")
                shared = _re.findall(r"<t[^>]*>([^<]*)</t>", sx)
            chunks = []
            for name in sorted(n for n in z.namelist() if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")):
                xml = z.read(name).decode("utf-8", errors="replace")
                # Inline strings
                inline = _re.findall(r"<t[^>]*>([^<]*)</t>", xml)
                # Shared-string refs: <c t="s"><v>IDX</v></c>
                refs = _re.findall(r'<c[^>]*t="s"[^>]*><v>(\d+)</v>', xml)
                resolved = [shared[int(i)] if int(i) < len(shared) else "" for i in refs]
                chunks.append("\n".join(filter(None, inline + resolved)))
            return "\n\n".join(chunks).strip() or None
    except Exception as e:
        _log.error(f"xlsx stdlib read failed: {path}: {e}")
        return None


def _read_epub(path: Path) -> str | None:
    """Extract text from .epub — zip of xhtml files, strip HTML from each."""
    import zipfile
    try:
        with zipfile.ZipFile(str(path)) as z:
            xhtml_names = sorted(n for n in z.namelist() if n.lower().endswith((".xhtml", ".html", ".htm")))
            chunks = []
            for name in xhtml_names:
                try:
                    html = z.read(name).decode("utf-8", errors="replace")
                    text = _strip_html(html)
                    if text:
                        chunks.append(text)
                except Exception:
                    pass
            return "\n\n---\n\n".join(chunks).strip() or None
    except Exception as e:
        _log.error(f"epub read failed: {path}: {e}")
        return None


def _read_rtf(path: Path) -> str | None:
    """Very rough RTF strip: remove control words and groups."""
    import re as _re
    try:
        raw = path.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        _log.error(f"rtf read failed: {path}: {e}")
        return None
    # Remove RTF headers, control words, and braces
    raw = _re.sub(r"\\[a-zA-Z]+-?\d* ?", "", raw)   # \b, \par0, \fs24 , etc.
    raw = _re.sub(r"\\'[0-9a-fA-F]{2}", "", raw)    # hex-escaped bytes
    raw = _re.sub(r"[{}]", "", raw)
    raw = _re.sub(r"\s+", " ", raw).strip()
    return raw or None


def _read_odt(path: Path) -> str | None:
    """Extract text from .odt (OpenOffice) — zip with content.xml."""
    import zipfile, re as _re
    try:
        with zipfile.ZipFile(str(path)) as z:
            xml = z.read("content.xml").decode("utf-8", errors="replace")
        # <text:p>, <text:span> etc. — strip tags
        xml = _re.sub(r"</text:p>", "\n", xml)
        xml = _re.sub(r"<[^>]+>", "", xml)
        from html import unescape
        return unescape(xml).strip() or None
    except Exception as e:
        _log.error(f"odt read failed: {path}: {e}")
        return None


def _read_ipynb(path: Path) -> str | None:
    """Extract markdown + code cells from a Jupyter notebook."""
    import json
    try:
        nb = json.loads(path.read_text(encoding="utf-8", errors="replace"))
    except Exception as e:
        _log.error(f"ipynb parse failed: {path}: {e}")
        return None
    out = []
    for cell in nb.get("cells", []):
        src = cell.get("source", "")
        if isinstance(src, list):
            src = "".join(src)
        if cell.get("cell_type") == "markdown":
            out.append(src)
        elif cell.get("cell_type") == "code":
            out.append("```python\n" + src + "\n```")
    return "\n\n".join(out).strip() or None


def _read_file(path: Path) -> str | None:
    """Read file content. Supports text, PDF, images, Office, ebooks.

    Strategy:
      1. MarkItDown as primary converter for rich formats (PDF/Office/HTML/EPUB/images).
         MarkItDown pulls in its own dependencies (pdfminer.six, python-docx, pptx, openpyxl…)
         and produces clean markdown that preserves tables, headings, and reading order.
      2. Stdlib / pypdf fallbacks when markitdown is unavailable or fails on a given file.
      3. Plain text read for anything that's just a text-like file.
    """
    ext = path.suffix.lower()

    # Images: use the vision-based describer (markitdown can do OCR too but we
    # already have a tuned vision pipeline with rate-limiting + describe_image).
    if ext in IMAGE_EXTENSIONS:
        try:
            return _describe_image(path)
        except Exception as e:
            _log.error(f"image read failed: {path}: {e}")
            return None

    # Formats where MarkItDown is noticeably better than our stdlib readers.
    MD_PRIMARY = {".pdf", ".docx", ".pptx", ".xlsx", ".epub",
                  ".html", ".htm", ".xhtml", ".rtf", ".odt",
                  ".csv", ".tsv", ".ipynb"}
    if ext in MD_PRIMARY:
        md_text = _markitdown_convert(path)
        if md_text:
            return md_text
        # Fall through to stdlib fallbacks below
        _log.info(f"markitdown returned no text for {path.name}, trying stdlib reader")

    # Stdlib fallbacks (work even without markitdown installed)
    if ext == ".pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(str(path))
            return "\n\n".join(page.extract_text() or "" for page in reader.pages)
        except ImportError:
            _log.warning("pypdf not installed — pip install pypdf to index PDFs")
            return None
        except Exception as e:
            _log.error(f"PDF read failed: {path}: {e}")
            return None
    if ext == ".docx":
        return _read_docx(path)
    if ext == ".pptx":
        return _read_pptx(path)
    if ext == ".xlsx":
        return _read_xlsx(path)
    if ext == ".epub":
        return _read_epub(path)
    if ext == ".rtf":
        return _read_rtf(path)
    if ext == ".odt":
        return _read_odt(path)
    if ext == ".ipynb":
        return _read_ipynb(path)
    if ext in (".html", ".htm", ".xhtml"):
        try:
            return _strip_html(path.read_text(encoding="utf-8", errors="replace"))
        except Exception as e:
            _log.error(f"html read failed: {path}: {e}")
            return None
    # Plain text
    try:
        return path.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        _log.error(f"read failed: {path}: {e}")
        return None


def index_url(url: str, tags: list[str] | None = None) -> dict:
    """Fetch a URL, convert to markdown, and index it.

    MarkItDown handles the full pipeline: HTTP fetch, content-type detection,
    HTML→markdown (preserving headings/tables/lists), PDF extraction, even
    YouTube transcripts. Falls back to urllib + stdlib HTML strip if
    MarkItDown fails (e.g. auth-walled page, unusual MIME).
    """
    import uuid as _uuid
    from urllib.parse import urlparse

    if not url.startswith(URL_SCHEMES):
        return {"url": url, "chunks": 0, "status": "invalid URL scheme"}

    uploads = Path(config.UPLOADS_DIR) / "kb"
    uploads.mkdir(parents=True, exist_ok=True)
    slug = _uuid.uuid4().hex[:8]
    parsed = urlparse(url)
    base = (parsed.netloc + parsed.path).rstrip("/").replace("/", "_")[-60:] or "page"

    # Primary path: MarkItDown converts the URL directly to markdown.
    md_text = _markitdown_convert(url)
    if md_text:
        dest = uploads / f"{slug}_{base}.md"
        header = (
            f"<!-- Source: {url} -->\n"
            f"<!-- Fetched: {time.strftime('%Y-%m-%d %H:%M:%S')} -->\n"
            f"<!-- Converter: markitdown -->\n\n"
        )
        dest.write_text(header + md_text, encoding="utf-8")
        all_tags = list(tags or []) + ["source:url", url]
        result = index_file(str(dest), tags=all_tags)
        result["url"] = url
        result["converter"] = "markitdown"
        return result

    # Fallback: raw urllib + stdlib HTML strip.
    import urllib.request
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (compatible; qwe-qwe/kb; +https://github.com/deepfounder-ai/qwe-qwe)",
        })
        with urllib.request.urlopen(req, timeout=20) as resp:
            ctype = resp.headers.get("content-type", "").split(";")[0].strip().lower()
            raw = resp.read()
    except Exception as e:
        _log.error(f"URL fetch failed: {url}: {e}")
        return {"url": url, "chunks": 0, "status": f"fetch failed: {e}"}

    if ctype.startswith("text/html") or ctype == "" or url.lower().endswith((".html", ".htm")):
        text = _strip_html(raw.decode("utf-8", errors="replace"))
        dest = uploads / f"{slug}_{base}.txt"
        dest.write_text(
            f"# Source: {url}\n# Fetched: {time.strftime('%Y-%m-%d %H:%M:%S')}\n\n{text}",
            encoding="utf-8",
        )
    elif ctype.startswith("application/pdf") or url.lower().endswith(".pdf"):
        dest = uploads / f"{slug}_{base}.pdf"
        dest.write_bytes(raw)
    elif ctype.startswith("text/"):
        text = raw.decode("utf-8", errors="replace")
        dest = uploads / f"{slug}_{base}.txt"
        dest.write_text(f"# Source: {url}\n\n{text}", encoding="utf-8")
    else:
        return {"url": url, "chunks": 0, "status": f"unsupported content-type: {ctype}"}

    all_tags = list(tags or []) + ["source:url", url]
    result = index_file(str(dest), tags=all_tags)
    result["url"] = url
    result["converter"] = "stdlib-fallback"
    return result


def index_file(filepath: str, tags: list[str] | None = None) -> dict:
    """Index a file into the unified memory collection (tag=knowledge).
    File content is auto-chunked by memory.save() and queued for synthesis."""
    import memory

    path = Path(filepath).expanduser().resolve()
    if not path.exists():
        return {"path": str(path), "chunks": 0, "status": "not found"}

    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS and ext != ".pdf" and ext not in IMAGE_EXTENSIONS:
        return {"path": str(path), "chunks": 0, "status": f"unsupported format: {ext}"}

    # Check if already indexed and unchanged
    mtime_key = f"rag:mtime:{path}"
    tags_key = f"rag:tags:{path}"
    stored_mtime = db.kv_get(mtime_key)
    current_mtime = str(path.stat().st_mtime)
    stored_tags = db.kv_get(tags_key) or ""
    new_tags = ",".join(tags) if tags else ""
    if stored_mtime == current_mtime and stored_tags == new_tags:
        return {"path": str(path), "chunks": 0, "status": "already up to date"}

    # Read file content
    content = _read_file(path)
    if not content or not content.strip():
        return {"path": str(path), "chunks": 0, "status": "empty file"}

    # Delete old chunks for this file (keeps kv entries — updated below on success)
    _delete_file_chunks(str(path))

    # Save via memory.save() — handles chunking, embeddings, FTS5, synthesis queue
    meta = {
        "source": str(path),
        "source_type": "file",
        "file_path": str(path),
        "filename": path.name,
    }
    if tags:
        meta["document_tags"] = tags

    memory.save(content, tag="knowledge", dedup=False, meta=meta)

    # Count chunks (for UI feedback) — matches memory.save() branching logic
    if len(content) > memory._CHUNK_THRESHOLD:
        chunk_count = len(memory._chunk_text(content))
    else:
        chunk_count = 1

    # Store mtime + tags
    db.kv_set(mtime_key, current_mtime)
    if tags:
        db.kv_set(tags_key, ",".join(tags))
    else:
        db.execute("DELETE FROM kv WHERE key = ?", (tags_key,))
    _log.info(f"indexed {path} → memory collection: {chunk_count} chunks")
    return {"path": str(path), "chunks": chunk_count, "status": "indexed"}


def index_directory(dirpath: str, recursive: bool = True) -> list[dict]:
    """Index all supported files in a directory."""
    path = Path(dirpath).expanduser().resolve()
    if not path.is_dir():
        return [{"path": str(path), "chunks": 0, "status": "not a directory"}]

    results = []
    pattern = "**/*" if recursive else "*"
    for f in sorted(path.glob(pattern)):
        if not f.is_file():
            continue
        ext = f.suffix.lower()
        if ext in ALL_INDEXABLE:
            result = index_file(str(f))
            results.append(result)
    return results


def search(query: str, limit: int = 5, tags: list[str] | None = None) -> list[dict]:
    """Search indexed files via unified memory collection.

    Args:
        query: search text.
        limit: max results.
        tags: optional tags filter (matches document_tags in payload).

    Returns [{text, file_path, chunk_index, score, tags}].
    """
    import memory

    # Search with tag=knowledge (files) via memory.search
    results = memory.search(query, limit=limit * 3, tag="knowledge")

    output = []
    for r in results:
        # Filter by document_tags if requested
        if tags:
            doc_tags = r.get("document_tags", [])
            if not any(t in doc_tags for t in tags):
                continue
        output.append({
            "text": r.get("text", ""),
            "file_path": r.get("file_path", r.get("source", "")),
            "chunk_index": r.get("chunk_index", 0),
            "score": round(r.get("score", 0.0), 4),
            "tags": r.get("document_tags", []),
        })
        if len(output) >= limit:
            break
    return output


def get_status() -> dict:
    """Get knowledge index status (unified memory collection)."""
    import memory
    from qdrant_client.models import Filter, FieldCondition, MatchValue

    count = 0
    try:
        qc = memory._get_qdrant()
        # Count knowledge-tagged points with source_type=file
        result = qc.count(
            config.QDRANT_COLLECTION,
            count_filter=Filter(must=[
                FieldCondition(key="source_type", match=MatchValue(value="file")),
            ]),
        )
        count = result.count
    except Exception:
        pass

    # Count indexed files from kv
    files = db.fetchone("SELECT COUNT(*) FROM kv WHERE key LIKE 'rag:mtime:%'")[0]

    return {"files": files, "chunks": count}


def stats() -> dict:
    """Alias for get_status() — for cli/telegram_bot compatibility."""
    s = get_status()
    return {"total_files": s["files"], "total_chunks": s["chunks"]}


def _delete_file_chunks(file_path: str):
    """Remove all chunks for a given file path from unified memory collection."""
    import memory
    from qdrant_client.models import Filter, FieldCondition, MatchValue, FilterSelector
    try:
        qc = memory._get_qdrant()
        qc.delete(
            config.QDRANT_COLLECTION,
            points_selector=FilterSelector(
                filter=Filter(must=[
                    FieldCondition(key="file_path", match=MatchValue(value=file_path)),
                ])
            ),
        )
    except Exception as e:
        _log.debug(f"delete chunks from memory collection failed: {e}")
    # Legacy RAG collection cleanup
    try:
        qc = _get_qdrant()
        qc.delete(
            RAG_COLLECTION,
            points_selector=FilterSelector(
                filter=Filter(must=[
                    FieldCondition(key="file_path", match=MatchValue(value=file_path)),
                ])
            ),
        )
    except Exception:
        pass
    db.fts_delete_match("fts_rag", "file_path", file_path)
    # fts_memory schema: (point_id, tag, text) — no file_path column.
    # Stale rows for re-indexed files remain but new chunks dedup by content in memory.save().


# ---------------------------------------------------------------------------
# Vision / image description
# ---------------------------------------------------------------------------

_vision_lock = threading.Lock()
_last_vision_call = 0.0


def _describe_image(path_or_bytes, prompt=None) -> str:
    """Describe an image using LM Studio vision API.

    Args:
        path_or_bytes: Path object or raw bytes of an image.
        prompt: optional text prompt for the vision model.

    Returns:
        Description string, or fallback placeholder on error.
    """
    import base64
    global _last_vision_call

    try:
        from PIL import Image
        import io
    except ImportError:
        _log.warning("Pillow not installed — pip install Pillow to process images")
        return "[image: description unavailable — Pillow not installed]"

    try:
        # Read bytes
        if isinstance(path_or_bytes, (str, Path)):
            img_bytes = Path(path_or_bytes).read_bytes()
        else:
            img_bytes = path_or_bytes

        # Resize to max 512px
        img = Image.open(io.BytesIO(img_bytes))
        img.thumbnail((512, 512))
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=85)
        resized_bytes = buf.getvalue()

        b64 = base64.b64encode(resized_bytes).decode("ascii")

        # Rate limit (thread-safe)
        with _vision_lock:
            elapsed = time.time() - _last_vision_call
            if elapsed < VISION_RATE_LIMIT:
                time.sleep(VISION_RATE_LIMIT - elapsed)
            _last_vision_call = time.time()

        import providers
        client = providers.get_client()
        model = providers.get_model()

        resp = client.chat.completions.create(
            model=model,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt or "Describe this image in detail. Include any text, diagrams, or visual elements you can see."},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
                ],
            }],
            max_tokens=512,
        )
        return resp.choices[0].message.content or "[image: empty response]"
    except Exception as e:
        _log.error(f"vision describe failed: {e}")
        return "[image: description unavailable]"


# ---------------------------------------------------------------------------
# PDF with vision fallback for scanned pages
# ---------------------------------------------------------------------------


def _read_pdf_with_vision(path) -> str:
    """Read PDF with page markers; notes scanned pages that lack text.

    Args:
        path: Path to the PDF file.

    Returns:
        Concatenated page text with markers.
    """
    try:
        from pypdf import PdfReader
    except ImportError:
        _log.warning("pypdf not installed — pip install pypdf to read PDFs")
        return ""

    path = Path(path)
    reader = PdfReader(str(path))
    parts = []
    for i, page in enumerate(reader.pages, 1):
        text = page.extract_text() or ""
        if len(text.strip()) < SCANNED_PAGE_THRESHOLD:
            parts.append(f"--- page {i} ---\n[scanned page {i} - text extraction not available]")
        else:
            parts.append(f"--- page {i} ---\n{text}")
    return "\n\n".join(parts)


# ---------------------------------------------------------------------------
# Scan path — analyse files before indexing
# ---------------------------------------------------------------------------


def scan_path(path_str: str, recursive: bool = True) -> dict:
    """Scan a file or directory and return indexing plan.

    Args:
        path_str: path to a file or directory.
        recursive: whether to recurse into subdirectories (default True).

    Returns:
        dict with ``files``, ``summary``, and ``estimate`` keys.
    """
    path = Path(path_str).expanduser().resolve()
    if not path.exists():
        return {"files": [], "summary": {}, "estimate": {}, "error": f"path not found: {path}"}

    # Gather file list
    if path.is_file():
        file_list = [path]
    else:
        pattern = "**/*" if recursive else "*"
        file_list = sorted(f for f in path.glob(pattern) if f.is_file())

    summary = {"text": 0, "pdf": 0, "image": 0, "unsupported": 0, "skipped": 0}
    est_chunks = 0
    est_time_cpu = 0.0
    est_time_gpu = 0.0
    gpu_required = False
    files = []

    for f in file_list[:MAX_SCAN_FILES]:
        ext = f.suffix.lower()
        size = f.stat().st_size

        if size > MAX_FILE_SIZE:
            files.append({"path": str(f), "type": "skipped", "reason": "too large", "size": size})
            summary["skipped"] += 1
            continue

        if ext in SUPPORTED_EXTENSIONS:
            ftype, method = "text", "chunk_text"
            summary["text"] += 1
            est_chunks += max(1, size // CHUNK_SIZE)
            est_time_cpu += 0.1
        elif ext == PDF_EXTENSION:
            ftype, method = "pdf", "pdf_extract"
            f_stat = f.stat()
            # Estimate pages from file size (~50KB per page) — avoids opening every PDF during scan
            est_pages = max(1, f_stat.st_size // 50000)
            summary["pdf"] += 1
            est_chunks += max(1, est_pages * 2)
            est_time_cpu += 0.2 * est_pages
            entry = {"path": str(f), "type": ftype, "method": method, "pages": est_pages, "size": size}
            files.append(entry)
            continue
        elif ext in IMAGE_EXTENSIONS:
            ftype, method = "image", "vision_describe"
            summary["image"] += 1
            est_chunks += 1
            est_time_gpu += 2.0
            gpu_required = True
        else:
            files.append({"path": str(f), "type": "unsupported", "ext": ext})
            summary["unsupported"] += 1
            continue

        files.append({"path": str(f), "type": ftype, "method": method, "size": size})

    return {
        "files": files,
        "summary": summary,
        "estimate": {
            "chunks": est_chunks,
            "time_cpu_sec": round(est_time_cpu, 1),
            "time_gpu_sec": round(est_time_gpu, 1),
            "gpu_required": gpu_required,
        },
    }


# ---------------------------------------------------------------------------
# Batch indexing with progress callbacks
# ---------------------------------------------------------------------------


def index_files_batch(files: list[dict], progress_cb=None, phase_cb=None,
                      tags: list[str] | None = None) -> list[dict]:
    """Batch-index a list of file descriptors from scan_path.

    Args:
        files: list of dicts with ``path``, ``type``, ``method`` keys.
        progress_cb: optional ``(current, total, path, phase, detail) -> None``.
        phase_cb: optional ``(event, count, estimate_sec) -> None``.
        tags: optional list of tag strings to attach to all indexed chunks.

    Returns:
        list of result dicts ``[{path, chunks, status, method}]``.
    """
    import memory

    # Normalize method names (frontend sends "text"/"pdf"/"vision",
    # backend uses "chunk_text"/"pdf_extract"/"vision_describe")
    _METHOD_MAP = {
        "text": "chunk_text", "pdf": "pdf_extract", "vision": "vision_describe",
        "chunk_text": "chunk_text", "pdf_extract": "pdf_extract",
        "vision_describe": "vision_describe", "pdf_scan": "pdf_scan",
    }
    for f in files:
        f["method"] = _METHOD_MAP.get(f.get("method", ""), "chunk_text")

    cpu_files = [f for f in files if f.get("method") in ("chunk_text", "pdf_extract")]
    gpu_files = [f for f in files if f.get("method") in ("vision_describe", "pdf_scan")]
    total = len(cpu_files) + len(gpu_files)
    results = []
    current = 0

    # Phase 1: CPU — text and normal PDF
    for f in cpu_files:
        current += 1
        fpath = f["path"]
        if progress_cb:
            progress_cb(current, total, fpath, "cpu", f.get("method", ""))
        result = index_file(fpath, tags=tags)
        result["method"] = f.get("method", "chunk_text")
        results.append(result)

    # Phase 2: GPU — images and scanned PDFs (vision-based extraction → memory.save)
    if gpu_files:
        est_sec = sum(2.0 if f.get("method") == "vision_describe" else 3.0 * f.get("pages", 1) for f in gpu_files)
        if phase_cb:
            phase_cb("gpu_warning", len(gpu_files), est_sec)

        for f in gpu_files:
            current += 1
            fpath = f["path"]
            method = f.get("method", "vision_describe")
            if progress_cb:
                progress_cb(current, total, fpath, "gpu", method)

            try:
                path = Path(fpath).expanduser().resolve()

                # Extract text via vision
                if method == "vision_describe":
                    content = _describe_image(path)
                elif method == "pdf_scan":
                    content = _read_pdf_with_vision(path)
                else:
                    content = _read_file(path)

                if not content or not content.strip():
                    results.append({"path": str(path), "chunks": 0, "status": "empty", "method": method})
                    continue

                # Save to unified memory collection (same path as CPU files)
                _delete_file_chunks(str(path))

                meta = {
                    "source": str(path),
                    "source_type": "file",
                    "file_path": str(path),
                    "filename": path.name,
                    "extraction_method": method,  # track how we got the text
                }
                if tags:
                    meta["document_tags"] = tags

                memory.save(content, tag="knowledge", dedup=False, meta=meta)

                # Count chunks for UI feedback
                chunk_count = len(memory._chunk_text(content)) if len(content) > memory._CHUNK_THRESHOLD else 1

                mtime_key = f"rag:mtime:{path}"
                tags_key = f"rag:tags:{path}"
                db.kv_set(mtime_key, str(path.stat().st_mtime))
                if tags:
                    db.kv_set(tags_key, ",".join(tags))
                else:
                    db.execute("DELETE FROM kv WHERE key = ?", (tags_key,))

                _log.info(f"indexed {path} → memory collection: {chunk_count} chunks (method={method})")
                results.append({"path": str(path), "chunks": chunk_count, "status": "indexed", "method": method})
            except Exception as e:
                _log.error(f"batch index failed for {fpath}: {e}")
                results.append({"path": fpath, "chunks": 0, "status": f"error: {e}", "method": method})

    return results


# ---------------------------------------------------------------------------
# List / delete indexed files
# ---------------------------------------------------------------------------


def list_indexed_files() -> list[dict]:
    """Return all indexed files enriched with chunk count + file size.

    Returns:
        list of ``{path, filename, tags, mtime, indexed_at, size, chunks}`` dicts.
        Chunks are counted by grouping Qdrant points by ``file_path`` in a single
        scroll — much cheaper than N separate count() calls.
    """
    rows = db.fetchall("SELECT key, value FROM kv WHERE key LIKE 'rag:mtime:%'")
    prefix_len = len("rag:mtime:")
    results = []
    paths = []
    for key, value in rows:
        fpath = key[prefix_len:]
        paths.append(fpath)
        tags_str = db.kv_get(f"rag:tags:{fpath}") or ""
        tags = [t.strip() for t in tags_str.split(",") if t.strip()] if tags_str else []
        p = Path(fpath)
        size = 0
        try:
            if p.exists():
                size = p.stat().st_size
        except Exception:
            pass
        results.append({
            "path": fpath,
            "filename": p.name,
            "name": p.name,
            "tags": tags,
            "mtime": float(value) if value else 0.0,
            "indexed_at": float(value) if value else 0.0,
            "size": size,
            "bytes": size,
            "chunks": 0,
        })
    # Bulk-count chunks per file_path across the unified memory collection
    try:
        import memory
        from qdrant_client.models import Filter, FieldCondition, MatchValue
        qc = memory._get_qdrant()
        counts: dict[str, int] = {}
        # Group by file_path — scroll through all knowledge-tagged points with file_path set
        offset = None
        limit = 512
        while True:
            points, offset = qc.scroll(
                config.QDRANT_COLLECTION,
                scroll_filter=Filter(must=[
                    FieldCondition(key="source_type", match=MatchValue(value="file")),
                ]),
                with_payload=["file_path"],
                with_vectors=False,
                limit=limit,
                offset=offset,
            )
            for p in points:
                fp = (p.payload or {}).get("file_path")
                if fp:
                    counts[fp] = counts.get(fp, 0) + 1
            if not offset:
                break
        for r in results:
            r["chunks"] = counts.get(r["path"], 0)
    except Exception as e:
        _log.debug(f"chunk count enrichment failed: {e}")
    return results


def delete_file(file_path: str) -> dict:
    """Delete a file from the RAG index.

    Args:
        file_path: absolute path of the file to remove.

    Returns:
        ``{path, status}`` dict.
    """
    path = Path(file_path).expanduser().resolve()
    _delete_file_chunks(str(path))
    # Remove mtime + tags tracking
    db.execute("DELETE FROM kv WHERE key = ?", (f"rag:mtime:{path}",))
    db.execute("DELETE FROM kv WHERE key = ?", (f"rag:tags:{path}",))
    return {"path": str(path), "status": "deleted"}
